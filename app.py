import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO

st.set_page_config(page_title="Анализатор курсов", layout="wide")
st.title("📊 Отчет по курсам (Регионы / Сертификаты)")

# Кэшируем чтение файла, чтобы не перечитывать его при каждом нажатии кнопок
@st.cache_data
def load_data(file):
    # Читаем все колонки сразу
    data = pd.read_excel(file, dtype=str)
    # Сразу чистим заголовки
    data.columns = [str(c).strip() for c in data.columns]
    return data

uploaded_file = st.file_uploader("Выберите файл Excel", type=["xlsx", "xls"])

if uploaded_file:
    # Показываем спиннер, пока файл читается
    with st.spinner('⏳ Обрабатываю большой файл... Пожалуйста, подождите (до 1-2 минут)'):
        try:
            df = load_data(uploaded_file)
            st.success('✅ Файл успешно загружен!')
        except Exception as e:
            st.error(f"Ошибка при чтении файла: {e}")
            st.stop()

    # Определяем нужные имена колонок
    col_id = "courses.id"
    col_region = "Область"
    col_cert = "Дата получения сертификата"
    col_name = "Наименование_курса"

    # Проверка на наличие колонок
    if col_id not in df.columns or col_region not in df.columns:
        st.error(f"В файле не найдены обязательные колонки: {col_id} или {col_region}")
        st.info(f"Доступные колонки: {', '.join(df.columns[:10])}...")
    else:
        # Интерфейс управления
        st.divider()
        mode = st.radio("Режим работы:", ["Все курсы", "По конкретному ID"])
        
        course_id_input = ""
        if mode == "По конкретному ID":
            course_id_input = st.text_input("Введите ID курса (например: 52)").strip()

        # Кнопка АНАЛИЗА появится только сейчас
        if st.button("📊 НАЧАТЬ АНАЛИЗ"):
            with st.spinner('Считаю статистику...'):
                current_course_name = "Все курсы"
                
                if mode == "По конкретному ID" and course_id_input:
                    filtered_df = df[df[col_id] == course_id_input].copy()
                    if not filtered_df.empty and col_name in filtered_df.columns:
                        current_course_name = filtered_df[col_name].iloc[0]
                    title = f"ОТЧЕТ ПО КУРСУ (ID: {course_id_input})"
                else:
                    filtered_df = df.copy()
                    title = "СВОДНЫЙ ОТЧЕТ ПО ВСЕМ КУРСАМ"

                if filtered_df.empty:
                    st.warning("Данные не найдены.")
                else:
                    # Чистим регионы
                    filtered_df[col_region] = filtered_df[col_region].str.strip().fillna("Не указано")
                    
                    # Логика сертификатов
                    filtered_df['has_cert'] = filtered_df[col_cert].notna() & (filtered_df[col_cert].astype(str).str.lower() != 'nan')
                    
                    report = filtered_df.groupby(col_region).agg(
                        total=(col_region, 'count'),
                        with_cert=('has_cert', 'sum')
                    ).reset_index()
                    
                    report['no_cert'] = report['total'] - report['with_cert']
                    report = report.sort_values(by='total', ascending=False)
                    
                    totals = (report['total'].sum(), report['with_cert'].sum(), report['no_cert'].sum())

                    # ВЫВОД РЕЗУЛЬТАТОВ
                    st.subheader(title)
                    if current_course_name != "Все курсы":
                        st.info(f"**Название:** {current_course_name}")
                    
                    # Метрики
                    m1, m2, m3 = st.columns(3)
                    m1.metric("Всего человек", totals[0])
                    m2.metric("С сертификатом", totals[1])
                    m3.metric("Без сертификата", totals[2])

                    # Таблица
                    st.dataframe(report, use_container_width=True)

                    # Подготовка Excel для скачивания
                    output = BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    ws.append([title, "", "", ""])
                    ws.append(["ID курса:", course_id_input if course_id_input else "Все", "", ""])
                    ws.append(["Название:", current_course_name, "", ""])
                    ws.append([])
                    ws.append(["Область", "Всего", "С сертификатом", "Без сертификата"])
                    
                    for _, row in report.iterrows():
                        ws.append([row['Область'], row['total'], row['with_cert'], row['no_cert']])
                    
                    ws.append(["ИТОГО", totals[0], totals[1], totals[2]])
                    for cell in ws[ws.max_row]: cell.font = Font(bold=True)
                    
                    wb.save(output)
                    
                    st.download_button(
                        label="💾 СКАЧАТЬ EXCEL ОТЧЕТ",
                        data=output.getvalue(),
                        file_name=f"report_{course_id_input or 'all'}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )